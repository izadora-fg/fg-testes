from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from copy import copy

import streamlit as st
import pandas as pd
import requests
import base64
import os

def enviar_github(nome_arquivo):

    token = st.secrets["GITHUB_TOKEN"]

    usuario = "izadora-fg"
    repositorio = "fg-testes"

    caminho = f"obras/{nome_arquivo}"

    with open(nome_arquivo, "rb") as f:
        conteudo = base64.b64encode(f.read()).decode()

    url = f"https://api.github.com/repos/{usuario}/{repositorio}/contents/{caminho}"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json"
    }

    # verifica se o arquivo já existe
    resposta = requests.get(url, headers=headers)

    sha = None

    if resposta.status_code == 200:
        sha = resposta.json()["sha"]

    dados = {
        "message": f"Atualiza {nome_arquivo}",
        "content": conteudo
    }

    if sha:
        dados["sha"] = sha

    resposta = requests.put(
        url,
        headers=headers,
        json=dados
    )

    if resposta.status_code not in (200, 201):
        st.error(f"Código: {resposta.status_code}")
        st.write(resposta.text)
        return False

    return True

    return resposta.status_code in (200, 201)

hoje = datetime.now(ZoneInfo("America/Sao_Paulo")).date()

# ======================================
# Configuração da página
# ======================================

st.set_page_config(
    page_title="Organograma da Obra",
    layout="centered"
)


# ======================================
# Carregar banco de cidades e engenheiros
# ======================================

cidades = pd.read_excel(
    "cidades.xlsx"
)

engenheiros = pd.read_excel(
    "engenheiros.xlsx"
)


# ======================================
# Preparar estados
# ======================================

estados = sorted(
    cidades["UF"].dropna().unique()
)


# ======================================
# Controle de páginas
# ======================================

if "pagina" not in st.session_state:
    st.session_state.pagina = "cabecalho"

# ======================================
# CABEÇALHO
# ======================================

if st.button("Limpar sessão"):
    st.session_state.clear()
    st.rerun()

if st.session_state.pagina == "cabecalho":

    st.title("Organograma da Obra")

    st.subheader("Informações Gerais")

    data = st.date_input(
        "Data do preenchimento",
        value=hoje,
        key="data"
)

    obra = st.text_input(
        "Código da obra",
        key = "obra"
    )

    estado = st.selectbox(
        "Estado",
        estados,
        index=None,
        key = "estado"
    )

    cidades_disponiveis = cidades[
        cidades["UF"] == estado
    ]["Cidade"].dropna()


    cidade = st.selectbox(
        "Cidade",
        sorted(cidades_disponiveis),
        index=None,
        key = "cidade"
    )

    responsavel = st.selectbox(
    "Responsável da obra",
    sorted(engenheiros["responsaveis"].dropna()),
    index=None,
    key = "responsavel"
)

    if responsavel is not None:

        cargo = engenheiros.loc[
            engenheiros["responsaveis"] == responsavel,
            "cargo"
        ].iloc[0]

    else:

        cargo = ""

    st.divider()

    if st.button("Próxima seção ➜"):

        erros = []

        if obra.strip() == "":
            erros.append("Informe o código da obra.")

        if estado == None:
            erros.append("Selecione o estado.")

        if cidade == None:
            erros.append("Selecione a cidade.")

        if responsavel == None:
            erros.append("Selecione o responsável da obra.")

        if erros:

            mensagem = "Preencha os campos abaixo antes de continuar:\n\n"

            for erro in erros:
                mensagem += f"- {erro}\n"

            st.warning(mensagem)

        else:

            st.session_state["dados_obra"] = obra
            st.session_state["dados_data"] = data
            st.session_state["dados_cidade"] = cidade
            st.session_state["dados_estado"] = estado
            st.session_state["dados_responsavel"] = responsavel
            st.session_state["dados_cargo"] = cargo

            st.session_state.pagina = "equipes"

            st.rerun()

# ======================================
# EQUIPES
# ======================================

elif st.session_state.pagina == "equipes":

    st.title("Seção 1 - Equipes")

    st.write("Informe todas as equipes que atuarão na obra.")

    quantidade_equipes = st.number_input(
        "Quantas equipes existem?",
        min_value=0,
        step=1,
        value=0,
        key="quantidade_equipes"
    )

    st.divider()

    for i in range(quantidade_equipes):

        st.subheader(f"Equipe {i+1}")

        funcao = st.text_input(
            "Função da equipe",
            key=f"funcao_{i}"
        )

        quantidade_funcionarios = st.number_input(
            "Quantidade de funcionários",
            min_value=1,
            step=1,
            value=1,
            key=f"qtd_func_{i}"
        )

        st.write("### Funcionários")

        for funcionario in range(quantidade_funcionarios):

            st.markdown(f"**Funcionário {funcionario + 1}**")

            col1, col2 = st.columns([1, 2])

            with col1:
                cargo = st.text_input(
                    "Cargo",
                    key=f"cargo_{i}_{funcionario}"
                )

            with col2:
                nome = st.text_input(
                    "Nome",
                    key=f"nome_{i}_{funcionario}"
                )

        st.divider()

    col1, col2 = st.columns(2)

    with col1:

        if st.button("⬅ Voltar", key="bt_vt_eq"):
            st.session_state.pagina = "cabecalho"
            st.rerun()

    with col2:

        avancar = st.button("Próxima seção ➜", key="bt_px_eq")

    if avancar:

        erros = []

        for i in range(quantidade_equipes):

            if st.session_state[f"funcao_{i}"].strip() == "":
                erros.append(f"Informe a função da Equipe {i+1}.")

            quantidade_funcionarios = st.session_state[f"qtd_func_{i}"]

            for funcionario in range(quantidade_funcionarios):

                if st.session_state[f"cargo_{i}_{funcionario}"].strip() == "":
                    erros.append(
                        f"Informe o cargo do Funcionário {funcionario+1} da Equipe {i+1}."
                    )

                if st.session_state[f"nome_{i}_{funcionario}"].strip() == "":
                    erros.append(
                        f"Informe o nome do Funcionário {funcionario+1} da Equipe {i+1}."
                    )

        if erros:

            mensagem = "Preencha os campos abaixo antes de continuar:\n\n"

            for erro in erros:
                mensagem += f"- {erro}\n"

            st.warning(mensagem)

        else:

            st.session_state["dados_equipes"] = {
                "quantidade_equipes": quantidade_equipes,
                "equipes": []
            }

            for i in range(quantidade_equipes):

                equipe = {
                    "funcao": st.session_state[f"funcao_{i}"],
                    "funcionarios": []
                }

                qtd = st.session_state[f"qtd_func_{i}"]

                for funcionario in range(qtd):

                    equipe["funcionarios"].append({
                        "cargo": st.session_state[f"cargo_{i}_{funcionario}"],
                        "nome": st.session_state[f"nome_{i}_{funcionario}"]
                    })

                st.session_state["dados_equipes"]["equipes"].append(equipe)

            st.session_state.pagina = "veiculos"
            st.rerun()

elif st.session_state.pagina == "veiculos":

    st.title("Seção 2 - Veículos Leves")

    st.write("Informe os veículos leves na obra.")

    quantidade_veiculos = st.number_input(
        "Quantos veículos leves existem?",
        min_value=0,
        step=1,
        value=0,
        key="quantidade_veiculos_leves"
    )

    st.divider()

    for i in range(quantidade_veiculos):

        with st.container(border=True):

            st.subheader(f"Veículo Leve {i+1} de {quantidade_veiculos}")

            col1, col2 = st.columns(2)

            with col1:
                tipo = st.text_input(
                    "Tipo",
                    key=f"tipo_leve_{i}"
                )

            with col2:
                placa = st.text_input(
                    "Placa",
                    key=f"placa_leve_{i}"
                )

            col1, col2 = st.columns(2)

            with col1:
                responsavel = st.text_input(
                    "Responsável",
                    key=f"responsavel_leve_{i}"
                )

            with col2:
                cargo = st.text_input(
                    "Cargo / Setor",
                    key=f"cargo_setor_leve_{i}"
                )

            observacao = st.text_area(
                "Observação",
                key=f"obs_leve_{i}"
            )

            liberado = st.radio(
                "Veículo liberado?",
                ["Sim", "Não"],
                horizontal=True,
                key=f"liberado_leve_{i}"
            )

    col1, col2 = st.columns(2)

    with col1:
        if st.button("⬅ Voltar"):
            st.session_state.pagina = "equipes"
            st.rerun()

    with col2:
        avancar = st.button(
            "Próxima seção ➜",
            key="bt_px_veiculos"
    )

    if avancar:

        erros = []

        for i in range(st.session_state["quantidade_veiculos_leves"]):

            if st.session_state[f"tipo_leve_{i}"].strip() == "":
                erros.append(f"Informe o tipo do Veículo {i+1}.")

            if st.session_state[f"placa_leve_{i}"].strip() == "":
                erros.append(f"Informe a placa do Veículo {i+1}.")

            if st.session_state[f"responsavel_leve_{i}"].strip() == "":
                erros.append(f"Informe o responsável do Veículo {i+1}.")

            if st.session_state[f"cargo_setor_leve_{i}"].strip() == "":
                erros.append(f"Informe o cargo/setor do Veículo {i+1}.")

        if erros:

            mensagem = "Preencha os campos abaixo antes de continuar:\n\n"

            for erro in erros:
                mensagem += f"- {erro}\n"

            st.warning(mensagem)

        else:

            st.session_state["dados_veiculos_leves"] = {
            "quantidade": st.session_state["quantidade_veiculos_leves"],
            "veiculos": []
        }

            for i in range(st.session_state["quantidade_veiculos_leves"]):

                st.session_state["dados_veiculos_leves"]["veiculos"].append({
                    "tipo": st.session_state.get(f"tipo_leve_{i}", ""),
                    "placa": st.session_state.get(f"placa_leve_{i}", ""),
                    "responsavel": st.session_state.get(f"responsavel_leve_{i}", ""),
                    "cargo_setor": st.session_state.get(f"cargo_setor_leve_{i}", ""),
                    "observacao": st.session_state.get(f"obs_leve_{i}", ""),
                    "liberado": st.session_state.get(f"liberado_leve_{i}", "")
                })

        st.session_state.pagina = "veiculos_pesados"
        st.rerun()

elif st.session_state.pagina == "veiculos_pesados":

    st.title("Seção 3 - Veículos Pesados")

    st.write("Informe os veículos pesados utilizados na obra.")

    quantidade_veiculos = st.number_input(
        "Quantos veículos pesados existem?",
        min_value=0,
        step=1,
        value=0,
        key="quantidade_veiculos_pesados"
    )

    st.divider()

    for i in range(st.session_state["quantidade_veiculos_pesados"]):

        with st.container(border=True):

            st.subheader(f"Veículo Pesado {i+1}")

            col1, col2 = st.columns(2)

            with col1:
                tipo = st.text_input(
                    "Tipo",
                    key=f"tipo_pesado_{i}"
                )

            with col2:
                placa = st.text_input(
                    "Placa",
                    key=f"placa_pesado_{i}"
                )

            col1, col2 = st.columns(2)

            with col1:
                responsavel = st.text_input(
                    "Responsável",
                    key=f"responsavel_pesado_{i}"
                )

            with col2:
                proprietario = st.text_input(
                    "Empresa proprietária",
                    key=f"proprietario_pesado_{i}"
                )

            observacao = st.text_area(
                "Observação",
                key=f"obs_pesado_{i}"
            )

    col1, col2 = st.columns(2)

    with col1:
        if st.button("⬅ Voltar"):
            st.session_state.pagina = "veiculos"
            st.rerun()

    with col2:
        
        st.session_state["dados_veiculos_pesados"] = {
            "quantidade": st.session_state["quantidade_veiculos_pesados"],
            "veiculos": []
    }

        for i in range(st.session_state["quantidade_veiculos_pesados"]):

            st.session_state["dados_veiculos_pesados"]["veiculos"].append({
                "tipo": st.session_state.get(f"tipo_pesado_{i}", ""),
                "placa": st.session_state.get(f"placa_pesado_{i}", ""),
                "responsavel": st.session_state.get(f"responsavel_pesado_{i}", ""),
                "proprietario": st.session_state.get(f"proprietario_pesado_{i}", ""),
                "observacao": st.session_state.get(f"obs_pesado_{i}", "")
        })

        if st.button("Próxima seção ➜"):
            st.session_state.pagina = "contratacoes"
            st.rerun()

elif st.session_state.pagina == "contratacoes":

    st.title("Seção 4 - Contratações")

    st.write("Informe os funcionários e equipamentos contratados para a obra.")

    # ==========================
    # Funcionários
    # ==========================

    st.header("Funcionários")

    quantidade_cargos = st.number_input(
        "Quantos tipos de cargos foram contratados?",
        min_value=0,
        step=1,
        value=0
    )

    for i in range(quantidade_cargos):

        with st.container(border=True):

            st.subheader(f"Cargo {i+1}")

            col1, col2 = st.columns([2,1])

            with col1:

                cargo = st.text_input(
                    "Cargo",
                    key=f"cargo_contratado_{i}"
                )

            with col2:

                quantidade = st.number_input(
                    "Quantidade",
                    min_value=1,
                    step=1,
                    key=f"qtd_cargo_{i}"
                )

            observacao_funcionario = st.text_area(
                "Observação da contratação",
                key=f"obs_funcionario_{i}"
            )


    st.divider()


    # ==========================
    # Equipamentos
    # ==========================

    st.header("Equipamentos")

    quantidade_equipamentos = st.number_input(
        "Quantos tipos de equipamentos foram contratados?",
        min_value=0,
        step=1,
        value=0
    )

    for i in range(quantidade_equipamentos):

        with st.container(border=True):

            st.subheader(f"Equipamento {i+1}")

            col1, col2 = st.columns([2,1])

            with col1:

                equipamento = st.text_input(
                    "Equipamento",
                    key=f"equipamento_{i}"
                )

            with col2:

                quantidade = st.number_input(
                    "Quantidade",
                    min_value=1,
                    step=1,
                    key=f"qtd_equipamento_{i}"
                )

            observacao_equipamento = st.text_area(
                "Observação da contratação",
                key=f"obs_equipamento_{i}"
            )


    st.divider()

    col1, col2 = st.columns(2)

    with col1:
        if st.button("⬅ Voltar"):
            st.session_state.pagina = "veiculos_pesados"
            st.rerun()

    with col2:
        if st.button("Finalizar"):

            st.session_state["dados_contratacoes"] = {
            "funcionarios": [],
            "equipamentos": []
        }

            for i in range(quantidade_cargos):

                st.session_state["dados_contratacoes"]["funcionarios"].append({
                    "cargo": st.session_state.get(f"cargo_contratado_{i}", ""),
                    "quantidade": st.session_state.get(f"qtd_cargo_{i}", 0),
                    "observacao": st.session_state.get(f"obs_funcionario_{i}", "")
                })


            for i in range(quantidade_equipamentos):

                st.session_state["dados_contratacoes"]["equipamentos"].append({
                    "equipamento": st.session_state.get(f"equipamento_{i}", ""),
                    "quantidade": st.session_state.get(f"qtd_equipamento_{i}", 0),
                    "observacao": st.session_state.get(f"obs_equipamento_{i}", "")
                })

            wb = load_workbook("organograma_modelo.xlsx")
            ws = wb.active

            ws["G3"] = st.session_state["dados_obra"]
            ws["E3"] = st.session_state["dados_data"]
            ws["I3"] = st.session_state["dados_cidade"]
            ws["F4"] = st.session_state["dados_responsavel"]
            ws["F5"] = st.session_state["dados_cargo"]

            # ======================================
            # EQUIPES
            # ======================================

            linha = 9

            dados_equipes = st.session_state["dados_equipes"]

            ws["E7"] = dados_equipes["quantidade_equipes"]


            # quantidade de linhas necessárias para equipes
            linhas_necessarias = 0

            for equipe in dados_equipes["equipes"]:
                linhas_necessarias += 1
                linhas_necessarias += len(equipe["funcionarios"])


            # o modelo já possui uma linha de funcionário
            linhas_extras = linhas_necessarias - 1


            if linhas_extras > 0:

                ws.insert_rows(
                    idx=12,
                    amount=linhas_extras
                )


            # preencher equipes

            for equipe in dados_equipes["equipes"]:

                ws[f"C{linha}"] = equipe["funcao"]

                linha += 1

                for funcionario in equipe["funcionarios"]:

                    ws[f"C{linha}"] = funcionario["cargo"]

                    ws[f"E{linha}"] = funcionario["nome"]

                    linha += 1

            # ======================================
            # VEÍCULOS LEVES
            # ======================================

            dados_veiculos_leves = st.session_state["dados_veiculos_leves"]

            quantidade_veiculos_leves = dados_veiculos_leves["quantidade"]

            ws[f"E{linha}"] = quantidade_veiculos_leves


            # cria linhas extras para veículos
            linhas_extras = quantidade_veiculos_leves - 1

            if linhas_extras > 0:

                ws.insert_rows(
                    idx=linha + 4,
                    amount=linhas_extras
                )


            linha += 3


            for veiculo in dados_veiculos_leves["veiculos"]:

                ws[f"C{linha}"] = veiculo["tipo"]

                ws[f"D{linha}"] = veiculo["placa"]

                ws[f"E{linha}"] = veiculo["responsavel"]

                ws[f"F{linha}"] = veiculo["cargo_setor"]

                ws[f"G{linha}"] = veiculo["observacao"]

                ws[f"I{linha}"] = veiculo["liberado"]

                linha += 1

            # ======================================
            # VEÍCULOS PESADOS
            # ======================================

            dados_veiculos_pesados = st.session_state["dados_veiculos_pesados"]

            quantidade_veiculos_pesados = dados_veiculos_pesados["quantidade"]

            ws[f"E{linha}"] = quantidade_veiculos_pesados


            linhas_extras = quantidade_veiculos_pesados - 1

            if linhas_extras > 0:

                ws.insert_rows(
                    idx=linha + 4,
                    amount=linhas_extras
                )

            linha += 3

            for veiculo in dados_veiculos_pesados["veiculos"]:

                ws[f"C{linha}"] = veiculo["tipo"]

                ws[f"D{linha}"] = veiculo["placa"]

                ws[f"E{linha}"] = veiculo["responsavel"]

                ws[f"F{linha}"] = veiculo["proprietario"]

                ws[f"G{linha}"] = veiculo["observacao"]

                linha += 1

            # ======================================
            # CONTRATAÇÕES
            # ======================================

            dados_contratacoes = st.session_state["dados_contratacoes"]


            # ---------- Funcionários ----------

            quantidade_cargos = len(
                dados_contratacoes["funcionarios"]
            )

            ws[f"E{linha}"] = quantidade_cargos


            linhas_extras = quantidade_cargos - 1

            if linhas_extras > 0:

                ws.insert_rows(
                    idx=linha + 4,
                    amount=linhas_extras
                )

            linha += 3


            for funcionario in dados_contratacoes["funcionarios"]:

                ws[f"C{linha}"] = funcionario["cargo"]

                ws[f"E{linha}"] = funcionario["quantidade"]

                ws[f"G{linha}"] = funcionario["observacao"]

                linha += 1


            # ---------- Equipamentos ----------

            quantidade_equipamentos = len(
                dados_contratacoes["equipamentos"]
            )

            ws[f"E{linha}"] = quantidade_equipamentos


            linhas_extras = quantidade_equipamentos - 1

            if linhas_extras > 0:

                ws.insert_rows(
                    idx=linha + 4,
                    amount=linhas_extras
                )

            linha += 3


            for equipamento in dados_contratacoes["equipamentos"]:

                ws[f"C{linha}"] = equipamento["equipamento"]

                ws[f"E{linha}"] = equipamento["quantidade"]

                ws[f"G{linha}"] = equipamento["observacao"]

                linha += 1

            quantidade_cargos = st.session_state.get(
                "quantidade_cargos",
                0
            )

            ws[f"E{linha}"] = quantidade_cargos

            linha += 3


            for i in range(quantidade_cargos):

                ws[f"C{linha}"] = st.session_state.get(
                    f"cargo_contratado_{i}",
                    ""
                )

                ws[f"E{linha}"] = st.session_state.get(
                    f"qtd_cargo_{i}",
                    ""
                )

                ws[f"G{linha}"] = st.session_state.get(
                    f"obs_funcionario_{i}",
                    ""
                )

                linha += 1



            quantidade_equipamentos = st.session_state.get(
                "quantidade_equipamentos",
                0
            )

            ws[f"E{linha}"] = quantidade_equipamentos

            linha += 3


            for i in range(quantidade_equipamentos):

                ws[f"C{linha}"] = st.session_state.get(
                    f"equipamento_{i}",
                    ""
                )

                ws[f"E{linha}"] = st.session_state.get(
                    f"qtd_equipamento_{i}",
                    ""
                )

                ws[f"G{linha}"] = st.session_state.get(
                    f"obs_equipamento_{i}",
                    ""
                )

                linha += 1

            nome_arquivo = f'{st.session_state["dados_obra"]}.xlsx'
            wb.save(nome_arquivo)

            sucesso = enviar_github(nome_arquivo)

            if sucesso:
                st.success("Arquivo enviado para o GitHub!")
                os.remove(nome_arquivo)
            else:
                st.error("Erro ao enviar para o GitHub.")

            st.success("Formulário preenchido com sucesso!")